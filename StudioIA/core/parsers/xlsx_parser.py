"""
core/parsers/xlsx_parser.py
Estrazione da file Excel (.xlsx): contenuto delle celle foglio per foglio,
e immagini incorporate (grafici salvati come immagine, foto incollate, ecc).

Nota: per fogli enormi (decine di migliaia di righe) limitiamo il numero di
righe lette per foglio, per evitare che l'indicizzazione di un file anomalo
blocchi l'app per minuti. Questo è pensato per documenti di studio/lavoro,
non per database tabellari giganti.
"""

import openpyxl
from openpyxl.utils import get_column_letter

from core.models import ExtractedImage, ParsedDocument
from core.parsers.utils import guess_mime_type

MAX_ROWS_PER_SHEET = 5000


def _describe_anchor(anchor) -> str:
    """Costruisce una descrizione leggibile della posizione di un'immagine
    (es. "cella D2"), leggendo solo gli attributi numerici riga/colonna.

    Nota: NON usiamo str()/repr() sull'oggetto anchor di openpyxl: in alcune
    versioni della libreria la sua rappresentazione testuale solleva un
    TypeError interno (bug noto di openpyxl, non del nostro codice). Per
    essere robusti leggiamo direttamente le coordinate numeriche.
    """
    try:
        start = getattr(anchor, "_from", None)
        if start is not None:
            col = get_column_letter(start.col + 1)
            row = start.row + 1
            return f"cella {col}{row}"
    except Exception:
        pass
    return "posizione non determinata"


def extract(file_path: str) -> ParsedDocument:
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return ParsedDocument(source_file=file_path, error=f"Excel illeggibile: {e}")

    full_text_parts = []
    images = []

    for sheet in wb.worksheets:
        sheet_label = f"foglio '{sheet.title}'"
        rows_text = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=False)):
            if row_idx >= MAX_ROWS_PER_SHEET:
                rows_text.append(f"[... foglio troncato oltre {MAX_ROWS_PER_SHEET} righe ...]")
                break
            cells_text = []
            for cell in row:
                if cell.value is not None:
                    cells_text.append(f"{cell.coordinate}={cell.value}")
            if cells_text:
                rows_text.append(" | ".join(cells_text))

        sheet_text = "\n".join(rows_text)
        if sheet_text.strip():
            full_text_parts.append(f"[{sheet_label}]\n{sheet_text}")

        # Immagini incorporate nel foglio. openpyxl non offre un'API pubblica
        # stabile per leggerne i byte grezzi: usiamo l'attributo interno
        # `_data()`, protetto da try/except così un eventuale cambiamento
        # futuro della libreria non fa crashare l'intera scansione, al più
        # perdiamo le immagini di quel foglio.
        try:
            sheet_images = getattr(sheet, "_images", [])
        except Exception:
            sheet_images = []

        for img_idx, img in enumerate(sheet_images):
            try:
                data = img._data()
                anchor_desc = _describe_anchor(getattr(img, "anchor", None))
                images.append(
                    ExtractedImage(
                        source_file=file_path,
                        location_hint=f"{sheet_label}, immagine {img_idx + 1} ({anchor_desc})",
                        image_bytes=data,
                        mime_type=guess_mime_type(data),
                        nearby_text=sheet_text[:1500],
                    )
                )
            except Exception:
                continue

    return ParsedDocument(source_file=file_path, text="\n\n".join(full_text_parts), images=images)
